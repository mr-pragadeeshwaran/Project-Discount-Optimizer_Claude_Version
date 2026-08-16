"""
build_wave_kam_sheet.py — the Monday KAM handoff: governed cuts + Wave-N tests.

One operational workbook the key-account manager can execute directly:
  Section A — GOVERNED CUTS: the tracker's issued weekly moves (from
              execution_log_template.csv) — permanent glide steps, both
              engines agreed, all gates passed.
  Section B — WAVE TESTS: the Unlock Pipeline's wave cells — 2-3 week
              EXPERIMENTS with an explicit strike/revert rule, issued to
              LEARN each cell's true response, not as recommendations.

Every row carries the exact discount and price to set, the predicted weekly
units (point + uncertainty band from the category coefficient's CI), and the
strike level (the tracker's own 5% tolerance): two weekly misses => revert.

Also writes output/DISCOUNT_PLAN/wave<N>_issued.csv — the machine-readable
issue log the scorer reads when the next weekly export lands.

Run:  python -X utf8 scripts/reports/build_wave_kam_sheet.py [--wave 1]
Out:  output/WAVE<N>_KAM_SHEET_v<K>.xlsx   (versioned; never overwrites)
"""
import argparse
import datetime
import glob
import json
import os
import re
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, ROOT)
sys.path.insert(0, HERE)

import build_stage_workbook as bsw          # shared loader: waves, floors, holds

INK, BODY, MUTED = "0F172A", "334155", "64748B"
HEAD_FILL = PatternFill("solid", fgColor="1E293B")
CUT_FILL = PatternFill("solid", fgColor="FEF3C7")     # amber — permanent move
TEST_FILL = PatternFill("solid", fgColor="DBEAFE")    # blue — experiment
HAIR = Border(bottom=Side(style="hair", color="CBD5E1"))
TOL = 0.95                                            # tracker's 5% volume tolerance


def F(sz=10, bold=False, color=BODY, italic=False):
    return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)


def _next_out(wave):
    n_max = 0
    for f in glob.glob(os.path.join(ROOT, "output", f"WAVE{wave}_KAM_SHEET_v*.xlsx")):
        m = re.search(r"_v(\d+)\.xlsx$", f)
        if m:
            n_max = max(n_max, int(m.group(1)))
    return os.path.join(ROOT, "output", f"WAVE{wave}_KAM_SHEET_v{n_max + 1}.xlsx")


def _cat_beta(run):
    S = json.load(open(os.path.join(run, "plan", "plan_summary.json"), encoding="utf-8"))
    return {c: (float(m.get("beta_disc") or 0), float(m.get("se_disc") or 0))
            for c, m in S.get("models", {}).items() if m.get("ok")}


def _predict(r, to_disc, betas):
    """Predicted weekly units at to_disc: point + CI band from the category
    coefficient (beta ± 1.96·se). Falls back to the cell's marg_beta."""
    u_now = float(r.get("cur_units_wk") or 0)
    delta = float(r["cur_disc"]) - float(to_disc)           # ppt removed
    b, se = betas.get(str(r["category"]), (np.nan, np.nan))
    if not np.isfinite(b):
        b, se = float(r.get("marg_beta") or 0), 0.0
    point = u_now * np.exp(-b * delta)
    pess = u_now * np.exp(-(b + 1.96 * se) * delta)         # biggest plausible drop
    opt = u_now * np.exp(-(b - 1.96 * se) * delta)          # discount was pointless
    return u_now, point, min(pess, opt), max(pess, opt)


def _spend_wk(units, mrp, disc):
    return units * float(mrp) * float(disc) / 100.0


def main(wave_n=1):
    run = bsw._latest_run()
    d = bsw.load_joined(run)
    betas = _cat_beta(run)
    today = datetime.date.today()
    monday = today + datetime.timedelta(days=((7 - today.weekday()) % 7) or 7)

    # ── Section A: governed cuts, straight from the tracker's issued log ──
    gov_rows = []
    xl_path = os.path.join(ROOT, "output", "DISCOUNT_PLAN", "execution_log_template.csv")
    if os.path.exists(xl_path):
        xl = pd.read_csv(xl_path)
        xl = xl[xl.get("recommended_action", "").astype(str) == "cut"]
        gov = xl.merge(d, on="cell_id", how="left", suffixes=("", "_d"))
        for _, r in gov.iterrows():
            to_d = float(r["recommended_disc"])
            u_now, point, lo, hi = _predict(r, to_d, betas)
            gov_rows.append((r, to_d, u_now, point, lo, hi, "GOVERNED CUT"))

    # ── Section B: this wave's test cells ──
    wv = d[d["wave"] == f"Wave {wave_n}"].sort_values("slice_cost_mo", ascending=False)
    test_rows = []
    for _, r in wv.iterrows():
        gap = float(r["cur_disc"]) - float(r["floor"])
        to_d = float(r["floor"]) if gap <= 3.0 else float(r["cur_disc"]) - 3.0
        u_now, point, lo, hi = _predict(r, to_d, betas)
        test_rows.append((r, to_d, u_now, point, lo, hi, "WAVE TEST"))

    all_rows = gov_rows + test_rows
    out_xlsx = _next_out(wave_n)
    wb = Workbook(); wb.remove(wb.active)

    # ── Sheet 1: apply ──
    ws = wb.create_sheet(f"KAM — Apply Mon {monday:%d %b}")
    ws.cell(1, 1, f"Monday {monday:%d %b %Y} — {len(gov_rows)} governed cuts + "
                  f"{len(test_rows)} Wave-{wave_n} tests").font = F(14, True, INK)
    ws.cell(2, 1, "AMBER = permanent glide step (both engines agreed, all gates). "
                  "BLUE = 2-3 week EXPERIMENT — issued to learn the cell's response; "
                  "auto-revert rule on the Predictions sheet. Set the discount; the "
                  "price column shows what the shopper will see.").font = \
        Font(name="Arial", size=9, italic=True, color=MUTED)
    headers = ["Type", "product_id", "cell_id", "Product", "Pack", "City",
               "Disc % now", "SET TO %", "Price now Rs", "SET price Rs",
               "Est. spend saved Rs/wk", "Applied? (Y/N)", "Actual disc set", "Notes"]
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(hr, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for i, (r, to_d, u_now, point, lo, hi, typ) in enumerate(all_rows, hr + 1):
        mrp = float(r["mrp"])
        saved = _spend_wk(u_now, mrp, r["cur_disc"]) - _spend_wk(point, mrp, to_d)
        note = ("aggregated residual cities — if the console has no 'Others' "
                "region, apply on remaining cities & note it here"
                if str(r["city"]) == "Others" else "")
        vals = [typ, r["product_id"], r["cell_id"], str(r["title"])[:38],
                r.get("grammage", ""), r["city"],
                round(float(r["cur_disc"]), 1), round(to_d, 1),
                round(float(r["cur_price"]), 2), round(mrp * (1 - to_d / 100), 2),
                round(saved), "", "", note]
        for j, v in enumerate(vals, 1):
            c = ws.cell(i, j, v)
            c.font = F(9); c.border = HAIR
            if j == 1:
                c.fill = CUT_FILL if typ == "GOVERNED CUT" else TEST_FILL
                c.font = F(9, bold=True, color=INK)
            if j == 11:
                c.number_format = "#,##0"
    for j, w in enumerate([13, 10, 22, 30, 9, 12, 8, 8, 10, 10, 12, 9, 10, 34], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(hr + 1, 4).coordinate

    # ── Sheet 2: predictions & strikes ──
    ws2 = wb.create_sheet("Predictions & strikes")
    ws2.cell(1, 1, "What we expect, and exactly when we retreat").font = F(14, True, INK)
    ws2.cell(2, 1, f"Strike = actual weekly units below {TOL:.0%} of the point "
                   "prediction. TWO consecutive strikes => revert the cell to its "
                   "previous discount (the tracker's own kill-switch rule). Scoring "
                   "runs when the next weekly export lands (weekly step B / the "
                   "wave scorer).").font = Font(name="Arial", size=9, italic=True,
                                               color=MUTED)
    h2 = ["product_id", "cell_id", "City", "Type", "SET TO %",
          "Units/wk now", "Predicted units/wk", "Prediction band (CI)",
          "STRIKE below units/wk", "What CONFIRM means"]
    for j, h in enumerate(h2, 1):
        c = ws2.cell(4, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    log_rows = []
    for i, (r, to_d, u_now, point, lo, hi, typ) in enumerate(all_rows, 5):
        confirm = ("volume holds 2-3 weeks -> the trim is banked"
                   if typ == "GOVERNED CUT"
                   else "volume holds 2-3 weeks -> cell joins the bankable set; "
                        "a revert costs ~nothing and still teaches its response")
        vals = [r["product_id"], r["cell_id"], r["city"], typ, round(to_d, 1),
                round(u_now, 1), round(point, 1),
                f"{lo:,.0f} .. {hi:,.0f}", round(point * TOL, 1), confirm]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(i, j, v); c.font = F(9); c.border = HAIR
        log_rows.append({
            "issued": today.isoformat(), "apply_week": monday.isoformat(),
            "type": typ, "wave": wave_n if typ == "WAVE TEST" else "",
            "product_id": r["product_id"], "cell_id": r["cell_id"],
            "city": r["city"], "from_disc": round(float(r["cur_disc"]), 2),
            "to_disc": round(to_d, 2), "units_wk_now": round(u_now, 2),
            "pred_units_wk": round(point, 2),
            "strike_units_wk": round(point * TOL, 2)})
    for j, w in enumerate([10, 22, 12, 13, 8, 10, 11, 16, 12, 52], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w

    # ── Sheet 3: how to run ──
    ws3 = wb.create_sheet("How to run this")
    steps = [
        ("How to run Monday's sheet", 13, True),
        ("", 9, False),
        (f"1. Monday {monday:%d %b}: set each row's discount to 'SET TO %'. Mark "
         "Applied=Y and note any deviation in 'Actual disc set'.", 10, False),
        ("2. AMBER rows are permanent moves — next week's sheet gives their next "
         "glide step. BLUE rows are experiments — hold them at the test level.", 10, False),
        ("3. Each week: drop the new platform export into input_data/ and run the "
         "weekly scoring step. Every row is checked against its prediction.", 10, False),
        ("4. A cell that strikes twice (see Predictions sheet) reverts to its "
         "previous discount — no meetings needed, the rule decides.", 10, False),
        ("5. After 2-3 weeks: confirmed BLUE cells join the bankable set with "
         "receipts; Wave 2 goes live on the calendar in the Stage Report.", 10, False),
        ("", 9, False),
        ("Nothing in this sheet uses cost assumptions — only units, prices and "
         "discount spend. Predictions come from the same validated model the "
         "receipts in the Stage Report certify.", 9, True),
    ]
    for i, (t, sz, b) in enumerate(steps, 1):
        ws3.cell(i, 1, t).font = F(sz, bold=b, color=INK if b else BODY)
        ws3.cell(i, 1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.column_dimensions["A"].width = 100

    wb.save(out_xlsx)
    log_path = os.path.join(ROOT, "output", "DISCOUNT_PLAN", f"wave{wave_n}_issued.csv")
    pd.DataFrame(log_rows).to_csv(log_path, index=False)
    print(f"[wave-kam] {len(gov_rows)} governed cuts + {len(test_rows)} wave-{wave_n} "
          f"tests -> {out_xlsx}")
    print(f"[wave-kam] issue log for the scorer -> {log_path}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--wave", type=int, default=1)
    main(ap.parse_args().wave)
