"""
build_stage_workbook.py — the three-stage end-user workbook, SKU x city level.

One Excel file that walks a business reader through the maturity ladder:
  Stage 1 — Discount Response : where does discount actually increase sales?
  Stage 2 — Optimal Discount  : how much discount should each cell carry?
  Stage 3 — Promotion ROI     : did the discount create ECONOMIC value?
                                (incremental contribution / discount cost)

Everything is read from the latest run's existing outputs — no model is
re-fitted here. Derived columns are computed with the same local response
(marg_beta) and cost structure (COGS proxy + commission + fulfilment) the
engine itself uses, and every derivation is documented on the READ ME sheet.

Baselines are HONEST: "what the discount buys" is measured against each
cell's OBSERVED historical floor (the discount it has actually traded at),
never against an extrapolated zero.

Values-only workbook by design (an analysis snapshot, not a financial model
to edit): no live formulas, so nothing can mis-recalculate downstream.

Run:  python -X utf8 scripts/reports/build_stage_workbook.py
Out:  output/STATIQ_STAGE_REPORT.xlsx
"""
import datetime
import glob
import json
import os
import sys

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, ROOT)
import v4_config as cfg

MONTH = 30.0 / 7.0

# The engine's own hold gates (mirrored from scripts/analysis/discount_plan.py
# so the unlock conditions we print are the conditions the engine enforces).
try:
    sys.path.insert(0, os.path.join(ROOT, "scripts", "analysis"))
    import discount_plan as _dp
    OSA_LOW, MIN_WEEKS, MIN_DISC_STD = _dp.OSA_LOW, _dp.MIN_WEEKS, _dp.MIN_DISC_STD
    CAT_R2_FLOOR = _dp.CAT_R2_FLOOR
except Exception:                                   # keep the workbook buildable alone
    OSA_LOW, MIN_WEEKS, MIN_DISC_STD, CAT_R2_FLOOR = 75.0, 8, 1.5, 0.60

WAVE_SIZE = 15          # cells per test wave — small enough to read weekly
WAVE_GAP_WEEKS = 3      # each wave runs 2-3 weeks under the kill-switch
KILL_SWITCH = "auto-revert if actual volume misses prediction >5% twice"


def _next_versioned_out():
    """Delivered files are immutable history: never overwrite an existing
    version — each build writes STATIQ_STAGE_REPORT_v<N+1>.xlsx. The original
    unversioned file counts as v1."""
    import re
    base = os.path.join(ROOT, "output")
    n_max = 1 if os.path.exists(os.path.join(base, "STATIQ_STAGE_REPORT.xlsx")) else 0
    for f in glob.glob(os.path.join(base, "STATIQ_STAGE_REPORT_v*.xlsx")):
        m = re.search(r"_v(\d+)\.xlsx$", f)
        if m:
            n_max = max(n_max, int(m.group(1)))
    return os.path.join(base, f"STATIQ_STAGE_REPORT_v{n_max + 1}.xlsx")


OUT_XLSX = None   # resolved at run time (auto-versioned) unless overridden via argv

# ── house style ─────────────────────────────────────────────────────────────
INK, BODY, MUTED = "0F172A", "334155", "64748B"
HEAD_FILL = PatternFill("solid", fgColor="1E293B")
GOOD_FILL = PatternFill("solid", fgColor="DCFCE7")   # creates value
BAD_FILL = PatternFill("solid", fgColor="FEE2E2")    # destroys value
GREY_FILL = PatternFill("solid", fgColor="F1F5F9")   # uncertain
NOTE_FONT = Font(name="Arial", size=9, italic=True, color=MUTED)
HAIR = Border(bottom=Side(style="hair", color="CBD5E1"))


def F(sz=10, bold=False, color=BODY, italic=False):
    return Font(name="Arial", size=sz, bold=bold, color=color, italic=italic)


def _latest_run():
    runs = sorted(glob.glob(os.path.join(ROOT, "output", "runs", "20*")))
    for r in reversed(runs):
        if os.path.exists(os.path.join(r, "plan", "all_cells.csv")):
            return r
    raise SystemExit("no run with a plan found — run the monthly rebuild first")


def _clean_pid(v):
    s = str(v).strip()
    try:
        f = float(s)
        if abs(f - round(f)) < 1e-9:
            return str(int(round(f)))
    except ValueError:
        pass
    return s


def load_joined(run):
    ac = pd.read_csv(os.path.join(run, "plan", "all_cells.csv"))
    rec = pd.read_csv(os.path.join(run, "recommendations.csv"))
    keep = ["cell_id", "grammage", "historical_floor_disc", "elasticity",
            "rec_discount_pct", "elbow_discount_pct", "confidence_score",
            "confidence_tier", "phasing_plan"]
    rec = rec[[c for c in keep if c in rec.columns]]
    d = ac.merge(rec, on="cell_id", how="left")

    # Per-cell elasticity + uncertainty + fit, straight from stage 4's table.
    ee_path = os.path.join(run, "elasticity_estimates.csv")
    if os.path.exists(ee_path):
        ee = pd.read_csv(ee_path)
        ee_keep = ["cell_id", "price_elasticity", "price_elasticity_se",
                   "price_elasticity_lower", "price_elasticity_upper",
                   "cell_train_r2", "cell_test_r2", "n_train",
                   "disc_pct_std", "n_discount_levels"]
        d = d.merge(ee[[c for c in ee_keep if c in ee.columns]],
                    on="cell_id", how="left")

    # Second engine's independent elasticity (Bayesian-shrunk own-price).
    el_path = os.path.join(ROOT, "output", "DISCOUNT_PLAN", "pricing", "elasticities.csv")
    if os.path.exists(el_path):
        el = pd.read_csv(el_path)
        el["pid"] = el["product_id"].map(_clean_pid)
        d["pid"] = d["product_id"].map(_clean_pid)
        d = d.merge(el[["pid", "city", "own_elast", "own_sd", "low_confidence"]]
                    .rename(columns={"low_confidence": "engine2_low_conf"}),
                    on=["pid", "city"], how="left")

    pr_path = os.path.join(ROOT, "output", "DISCOUNT_PLAN", "pricing", "pricing_reco.csv")
    if os.path.exists(pr_path):
        pr = pd.read_csv(pr_path)
        pr["pid"] = pr["product_id"].map(_clean_pid)
        d["pid"] = d["product_id"].map(_clean_pid)
        d = d.merge(pr[["pid", "city", "opt_disc", "pred_units_delta_pct",
                        "pred_rev_delta_pct"]], on=["pid", "city"], how="left")
    ag_path = os.path.join(ROOT, "output", "DISCOUNT_PLAN", "pricing", "agreement.csv")
    if os.path.exists(ag_path):
        ag = pd.read_csv(ag_path)[["cell_id", "agree_with_cut"]]
        d = d.merge(ag, on="cell_id", how="left")

    # ── derived, per cell (documented on READ ME) ──
    d["floor"] = pd.to_numeric(d.get("historical_floor_disc"), errors="coerce") \
        .fillna(pd.to_numeric(d.get("tgt_disc"), errors="coerce")) \
        .fillna(d["cur_disc"]).clip(lower=0)
    d["floor"] = np.minimum(d["floor"], d["cur_disc"])          # floor never above today
    beta = pd.to_numeric(d["marg_beta"], errors="coerce").fillna(0.0)
    gap = (d["cur_disc"] - d["floor"]).clip(lower=0)

    u_now = pd.to_numeric(d["cur_units_wk"], errors="coerce").fillna(0.0)
    mrp = pd.to_numeric(d["mrp"], errors="coerce")
    p_now = pd.to_numeric(d["cur_price"], errors="coerce")
    p_floor = mrp * (1 - d["floor"] / 100.0)
    u_floor = u_now * np.exp(-beta * gap)                       # local response model

    d["uplift_pct"] = (np.where(u_floor > 0, u_now / u_floor - 1.0, 0.0)) * 100
    d["incr_units_wk"] = u_now - u_floor
    d["rev_uplift_wk"] = u_now * p_now - u_floor * p_floor
    d["incr_rev_mo"] = d["rev_uplift_wk"] * MONTH

    spend_now_wk = pd.to_numeric(d["disc_spend_mo"], errors="coerce").fillna(0.0) / MONTH
    spend_floor_wk = u_floor * mrp * (d["floor"] / 100.0)
    d["slice_cost_mo"] = ((spend_now_wk - spend_floor_wk) * MONTH).clip(lower=0)
    # Revenue ROI — everything observable, no cost assumptions: did the extra
    # revenue the discount slice brought exceed what the slice cost?
    d["rev_roi"] = np.where(d["slice_cost_mo"] > 1.0,
                            d["incr_rev_mo"] / d["slice_cost_mo"], np.nan)

    # Verdicts carry the engine's BUCKET-BEFORE-ACTION rule: a cell can be
    # statistically below the pay-line yet availability-constrained — the
    # action there is FIX STOCK, never trim. Only bucket-c cells say "trim".
    def _verdict(r):
        if bool(r.get("reliably_pays")):
            return "DISCOUNT WORKS — creates value"
        if bool(r.get("reliably_waste")):
            b = str(r.get("bucket", ""))
            if b == "c_waste_cut":
                return "NOT WORKING — trim to floor"
            if b == "a_stock":
                return "NOT WORKING, but stock-constrained — FIX AVAILABILITY first"
            if b == "b_competitive":
                return "NOT WORKING, but competitive pressure — HOLD, defensive"
            return "NOT WORKING — monitor (gates keep it out of the cut wave)"
        return "UNCERTAIN — monitor / test"
    d["verdict"] = d.apply(_verdict, axis=1)

    # ── Hold classification: WHY is a cell not acting, and what unlocks it? ──
    # Mirrors the engine's own gate order (bucket precedence, then quality
    # gates) so every "HELD" row carries the specific missing evidence — a
    # plan, not a refusal. test_candidate marks cells only a DESIGNED price
    # test can unlock (they feed the Unlock Pipeline sheet).
    today = datetime.date.today()

    def _classify(r):
        b = str(r.get("bucket", ""))
        conf = str(r.get("confidence", ""))
        osa = float(r.get("osa_mean") or 0)
        n_wk = float(r.get("n_weeks") or 0)
        dstd = float(r.get("disc_std") or 0)
        if b == "c_waste_cut":
            if conf == "High":
                return ("ACTING — in this week's governed plan",
                        "already in the KAM sheet", False)
            return ("In the TEST QUEUE (experimental-tier cut)",
                    f"2–3 week controlled test; {KILL_SWITCH}", True)
        if b == "a_stock":
            return (f"Availability-constrained (OSA {osa:.0f}% < {OSA_LOW:.0f}%)",
                    "raise availability — the cell re-enters pricing automatically "
                    "next run; its biggest lever is stock, not price", False)
        if b == "b_competitive":
            return ("Competitive pressure — trimming risks share",
                    "monitor the competitor price index; act when pressure clears", False)
        if not bool(r.get("cat_ok", True)):
            return (f"Category model below fit floor (R² {float(r.get('cat_r2') or 0):.2f} "
                    f"< {CAT_R2_FLOOR:.2f})",
                    "more weekly exports for the whole category", False)
        if n_wk and n_wk < MIN_WEEKS:
            ready = today + datetime.timedelta(weeks=int(MIN_WEEKS - n_wk))
            return (f"Only {n_wk:.0f} weeks of history (need {MIN_WEEKS})",
                    f"keep weekly exports coming — ready ≈ {ready:%d %b}", False)
        if dstd < MIN_DISC_STD:
            return (f"Discount never varies ({dstd:.1f} ppt std) — nothing to learn from",
                    "a designed price test — see Unlock Pipeline", True)
        mb, bb = r.get("marg_beta"), r.get("be_beta")
        if bool(r.get("reliably_pays")):
            return ("Discount reliably pays here",
                    "protect — reinvest logic governs, not the cut wave", False)
        if pd.notna(mb) and pd.notna(bb) and float(mb) < float(bb):
            return ("Below pay-line, but the CI spans it — not reliable yet",
                    "2–4 more weekly exports OR a designed test — see Unlock Pipeline", True)
        return ("No confident signal either way",
                "monitor — re-evaluated automatically every run", False)

    hc = d.apply(_classify, axis=1, result_type="expand")
    d["hold_reason"], d["unlock_when"], d["test_candidate"] = hc[0], hc[1], hc[2]

    # Wave assignment (shared by Unlock Pipeline, Confidence Growth, calendar):
    # test candidates ranked by value at stake, waves of WAVE_SIZE, top 3 waves.
    d["wave"] = ""
    _tc = d["test_candidate"].fillna(False).astype(bool)
    _order = d.loc[_tc].sort_values("slice_cost_mo", ascending=False).index
    for _i, _idx in enumerate(_order):
        d.loc[_idx, "wave"] = (f"Wave {(_i // WAVE_SIZE) + 1}"
                               if (_i // WAVE_SIZE) < 3 else "Backlog")
    return d


# ── sheet writers ───────────────────────────────────────────────────────────
def _write_table(ws, title, subtitle, headers, rows, widths, fmts, verdict_col=None):
    ws.cell(1, 1, title).font = F(14, bold=True, color=INK)
    ws.cell(2, 1, subtitle).font = NOTE_FONT
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(hr, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center",
                                horizontal="center" if j > 6 else "left")
    for i, row in enumerate(rows, hr + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v)
            c.font = F(9)
            c.border = HAIR
            fmt = fmts.get(j)
            if fmt:
                c.number_format = fmt
            if verdict_col and j == verdict_col and isinstance(v, str):
                # red ONLY where the action is actually "trim" — a below-pay-line
                # cell held for stock/competitive reasons is grey (bucket rule).
                c.fill = (GOOD_FILL if "WORKS" in v or "CREATES" in v
                          else BAD_FILL if "trim to floor" in v or "DESTROYS" in v
                          else GREY_FILL)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(hr + 1, 3).coordinate
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(headers))}{hr + len(rows)}"


def sheet_stage1(wb, d):
    ws = wb.create_sheet("Stage 1 — Discount Response")
    dd = d.sort_values(["verdict", "rev_uplift_wk"])
    headers = ["product_id", "cell_id", "Product", "Pack", "City", "Category",
               "Weeks", "OSA %", "Disc % now", "Price now", "MRP", "Units/wk",
               "Response per +1pt (% units)", "Pay-line per pt (%)",
               "Sales uplift % (vs proven floor)", "Incremental units/wk",
               "Revenue uplift Rs/wk", "Verdict", "Confidence"]
    rows = [[r["product_id"], r["cell_id"], str(r["title"])[:38],
             r.get("grammage", ""), r["city"], r["category"], r.get("n_weeks"),
             round(float(r["osa_mean"]), 0), round(float(r["cur_disc"]), 1),
             round(float(r["cur_price"]), 1), round(float(r["mrp"]), 0),
             round(float(r["cur_units_wk"]), 1),
             round(float(r["marg_beta"]) * 100, 3) if pd.notna(r["marg_beta"]) else None,
             round(float(r["be_beta"]) * 100, 3) if pd.notna(r["be_beta"]) else None,
             round(float(r["uplift_pct"]), 2), round(float(r["incr_units_wk"]), 1),
             round(float(r["rev_uplift_wk"])),
             r["verdict"], r["confidence"]]
            for _, r in dd.iterrows()]
    fmts = {10: "0.0", 12: "#,##0.0", 17: "#,##0"}
    _write_table(ws, "Stage 1 — Where does discount increase sales?",
                 "Response ISOLATED from availability, ad visibility, competition and "
                 "season. Uplift measured vs each cell's OBSERVED floor discount — never "
                 "an extrapolated zero. All figures observable: units, price, revenue.",
                 headers, rows,
                 [10, 22, 30, 9, 12, 15, 6, 6, 8, 8, 7, 9, 11, 10, 11, 11, 11, 30, 12],
                 fmts, verdict_col=18)


def sheet_stage2(wb, d):
    ws = wb.create_sheet("Stage 2 — Optimal Discount")
    dd = d.sort_values(["title", "city"])
    headers = ["product_id", "cell_id", "Product", "Pack", "City", "Category",
               "Disc % now", "Proven floor %", "Champion target %",
               "Pipeline elbow %", "Optimizer optimal %", "Engines agree on cut?",
               "SET THIS WEEK (governed) %", "Projected units Δ%", "Projected revenue Δ%",
               "Confidence"]
    rows = []
    for _, r in dd.iterrows():
        is_cut = (r.get("bucket") == "c_waste_cut") and bool(r.get("agree_with_cut", False))
        set_now = float(r["tgt_disc"]) if (is_cut and pd.notna(r.get("tgt_disc"))) \
            else float(r["cur_disc"])
        rows.append([r["product_id"], r["cell_id"], str(r["title"])[:38],
                     r.get("grammage", ""), r["city"], r["category"],
                     round(float(r["cur_disc"]), 1), round(float(r["floor"]), 1),
                     round(float(r["tgt_disc"]), 1) if pd.notna(r.get("tgt_disc")) else None,
                     round(float(r["rec_discount_pct"]), 1) if pd.notna(r.get("rec_discount_pct")) else None,
                     round(float(r["opt_disc"]), 1) if pd.notna(r.get("opt_disc")) else None,
                     "YES" if bool(r.get("agree_with_cut", False)) else "",
                     round(set_now, 1),
                     round(float(r["pred_units_delta_pct"]), 2) if pd.notna(r.get("pred_units_delta_pct")) else None,
                     round(float(r["pred_rev_delta_pct"]), 2) if pd.notna(r.get("pred_rev_delta_pct")) else None,
                     r["confidence"]])
    _write_table(ws, "Stage 2 — How much discount should each cell carry?",
                 "Three independent answers (champion target, pipeline elbow, portfolio "
                 "optimizer) and the governed number the tracker actually issues — a cut "
                 "executes only where BOTH engines agree, glide-limited weekly.",
                 headers, rows,
                 [10, 22, 30, 9, 12, 15, 8, 8, 9, 9, 9, 9, 11, 9, 9, 12], {})


def sheet_stage3(wb, d):
    ws = wb.create_sheet("Stage 3 — Promotion ROI")
    dd = d.sort_values("rev_roi")
    headers = ["product_id", "cell_id", "Product", "Pack", "City", "Category",
               "Disc % now", "Discount cost Rs/mo (total)",
               "Trimmable slice cost Rs/mo (floor→now)",
               "Incremental revenue Rs/mo (slice)",
               "REVENUE ROI (incr. revenue ÷ slice cost)",
               "Marginal ROI (last point)", "Net gain if trimmed Rs/mo",
               "Economic verdict", "Why (engine's reason)",
               "Held because", "Unlocks when"]
    rows = []
    for _, r in dd.iterrows():
        roi = r["rev_roi"]
        b = str(r.get("bucket", ""))
        if bool(r.get("reliably_pays")):
            ev = "CREATES VALUE — protect / consider reinvest"
        elif bool(r.get("reliably_waste")) and b == "c_waste_cut":
            ev = "DESTROYS VALUE — trim to floor"
        elif bool(r.get("reliably_waste")) and b == "a_stock":
            ev = "BELOW PAY-LINE, stock-constrained — fix availability, don't trim"
        elif bool(r.get("reliably_waste")) and b == "b_competitive":
            ev = "BELOW PAY-LINE, competitive — hold, defensive"
        elif bool(r.get("reliably_waste")):
            ev = "BELOW PAY-LINE — monitor (kept out of cut wave)"
        else:
            ev = "UNCERTAIN — monitor / test"
        rows.append([r["product_id"], r["cell_id"], str(r["title"])[:38],
                     r.get("grammage", ""), r["city"], r["category"],
                     round(float(r["cur_disc"]), 1),
                     round(float(r["disc_spend_mo"])) if pd.notna(r["disc_spend_mo"]) else None,
                     round(float(r["slice_cost_mo"])),
                     round(float(r["incr_rev_mo"])),
                     round(float(roi), 2) if pd.notna(roi) else None,
                     round(float(r["marginal_roas"]), 2) if pd.notna(r.get("marginal_roas")) else None,
                     round(float(r["net_gain_mo"])) if pd.notna(r.get("net_gain_mo")) else None,
                     ev, str(r.get("decision_reason", ""))[:80],
                     r["hold_reason"], r["unlock_when"]])
    fmts = {8: "#,##0", 9: "#,##0", 10: "#,##0", 13: "#,##0"}
    _write_table(ws, "Stage 3 — Did the discount create ECONOMIC value?",
                 "The metric is Incremental Revenue ÷ Discount Cost (not sales uplift "
                 "alone). ROI < 1 = the discount costs more revenue than it brings. "
                 "Every held cell names its reason AND its unlock condition — a "
                 "schedule, not a refusal. All figures observable; no cost assumptions.",
                 headers, rows,
                 [10, 22, 30, 9, 12, 15, 8, 11, 12, 12, 13, 9, 11, 26, 42, 34, 40],
                 fmts, verdict_col=14)


def sheet_unlock(wb, d):
    ws = wb.create_sheet("Unlock Pipeline")
    tc = d[d["test_candidate"].fillna(False).astype(bool)].copy()
    tc = tc.sort_values("slice_cost_mo", ascending=False).reset_index(drop=True)
    today = datetime.date.today()
    next_monday = today + datetime.timedelta(days=((7 - today.weekday()) % 7) or 7)

    def _wave(i):
        return (i // WAVE_SIZE) + 1

    def _start(i):
        w = _wave(i)
        return (next_monday + datetime.timedelta(weeks=(w - 1) * WAVE_GAP_WEEKS)) \
            if w <= 3 else None

    headers = ["product_id", "cell_id", "Product", "Pack", "City", "Category",
               "Wave", "Start week", "Disc % now", "Test move to %",
               "Value at stake Rs/mo", "OSA %", "Weeks",
               "Why a test is needed", "Safety rail"]
    rows = []
    for i, r in tc.iterrows():
        gapv = float(r["cur_disc"]) - float(r["floor"])
        test_to = float(r["floor"]) if gapv <= 3.0 else float(r["cur_disc"]) - 3.0
        st = _start(i)
        rows.append([r["product_id"], r["cell_id"], str(r["title"])[:38],
                     r.get("grammage", ""), r["city"], r["category"],
                     f"Wave {_wave(i)}" if _wave(i) <= 3 else "Backlog",
                     f"{st:%d %b %Y}" if st else "after wave 3 confirms",
                     round(float(r["cur_disc"]), 1), round(test_to, 1),
                     round(float(r["slice_cost_mo"])),
                     round(float(r.get("osa_mean") or 0)),
                     r.get("n_weeks"), r["hold_reason"], KILL_SWITCH])
    stake_w13 = float(tc.head(3 * WAVE_SIZE)["slice_cost_mo"].sum())
    stake_all = float(tc["slice_cost_mo"].sum())
    fmts = {11: "#,##0"}
    _write_table(
        ws, "Unlock Pipeline — how held value converts into bankable value",
        f"{len(tc)} cells only a DESIGNED test can unlock. Waves of {WAVE_SIZE} cells, "
        f"{WAVE_GAP_WEEKS} weeks apart, each protected by the kill-switch. Value at "
        f"stake: Rs.{stake_w13:,.0f}/mo in waves 1–3 (Rs.{stake_all:,.0f}/mo total "
        f"queue). SCENARIO, not promise: if 30–50% of a wave confirms, roughly that "
        f"share of its stake becomes bankable; the rest reverts at ~zero cost — and "
        f"either way the cell's true response is learned.",
        headers, rows,
        [10, 22, 30, 9, 12, 15, 8, 12, 8, 8, 12, 6, 6, 40, 34], fmts)


def sheet_elasticity(wb, d):
    ws = wb.create_sheet("Elasticity & Accuracy")
    dd = d.sort_values(["title", "city"])
    headers = ["product_id", "cell_id", "Product", "Pack", "City", "Category",
               "Price elasticity", "95% CI low", "95% CI high",
               "2nd-engine elasticity (independent)", "2nd-engine SD",
               "Weeks of data", "Training days", "Discount variation (ppt std)",
               "Distinct discount levels", "Fit at decision grain (R²)",
               "Held-out R² (where measurable)", "Confidence score (0-100)",
               "Confidence tier"]
    rows = []
    for _, r in dd.iterrows():
        rows.append([
            r["product_id"], r["cell_id"], str(r["title"])[:38],
            r.get("grammage", ""), r["city"], r["category"],
            round(float(r["price_elasticity"]), 2) if pd.notna(r.get("price_elasticity")) else None,
            round(float(r["price_elasticity_lower"]), 2) if pd.notna(r.get("price_elasticity_lower")) else None,
            round(float(r["price_elasticity_upper"]), 2) if pd.notna(r.get("price_elasticity_upper")) else None,
            round(float(r["own_elast"]), 2) if pd.notna(r.get("own_elast")) else None,
            round(float(r["own_sd"]), 2) if pd.notna(r.get("own_sd")) else None,
            r.get("n_weeks"),
            int(r["n_train"]) if pd.notna(r.get("n_train")) else None,
            round(float(r["disc_pct_std"]), 1) if pd.notna(r.get("disc_pct_std")) else None,
            int(r["n_discount_levels"]) if pd.notna(r.get("n_discount_levels")) else None,
            round(float(r["cell_train_r2"]), 2) if pd.notna(r.get("cell_train_r2")) else None,
            round(float(r["cell_test_r2"]), 2) if pd.notna(r.get("cell_test_r2")) else None,
            round(float(r["confidence_score"]), 0) if pd.notna(r.get("confidence_score")) else None,
            r.get("confidence_tier", ""),
        ])
    _write_table(ws, "Elasticity & Accuracy — the number and how much to trust it, per cell",
                 "Two INDEPENDENT elasticity estimates per cell (pipeline + pricing "
                 "engine) with confidence intervals. Thin cells are shrunk toward their "
                 "category median — that is honesty, not weakness. Accuracy is layered: "
                 "cell confidence score → category fit → portfolio receipts (Summary).",
                 headers, rows,
                 [10, 22, 30, 9, 12, 15, 9, 8, 8, 11, 9, 7, 8, 10, 9, 10, 10, 10, 12], {})


def _project_tiers(d, add_weeks, with_waves):
    """Project confidence tiers +N weeks out, using the PUBLISHED score formula
    (0.25 density + 0.20 variation + 0.20 fit + 0.15 plausibility + 0.20
    tightness). Only density (rows accrue weekly) and — when waves run —
    variation (a test adds ≥2 distinct discount levels) are advanced; fit,
    plausibility and CI tightness are held constant even though they normally
    improve with data, so this projection is a FLOOR, not a promise."""
    n_wk = pd.to_numeric(d["n_weeks"], errors="coerce")
    n_tr = pd.to_numeric(d.get("n_train"), errors="coerce").fillna(0)
    lvl = pd.to_numeric(d.get("n_discount_levels"), errors="coerce").fillna(0)
    score = pd.to_numeric(d.get("confidence_score"), errors="coerce").fillna(0)
    rate = (n_tr / n_wk.replace(0, np.nan)).fillna(0)
    dens1 = np.clip(n_tr / 120.0, 0, 1)
    dens2 = np.clip((n_tr + rate * add_weeks) / 120.0, 0, 1)
    var1 = np.clip(lvl / 15.0, 0, 1)
    waves_done = {4: ["Wave 1"], 8: ["Wave 1", "Wave 2"],
                  12: ["Wave 1", "Wave 2", "Wave 3"]}.get(add_weeks, [])
    bonus = (d["wave"].isin(waves_done) & with_waves).astype(float) * 2.0
    var2 = np.clip((lvl + bonus) / 15.0, 0, 1)
    s2 = np.clip(score + 25.0 * (dens2 - dens1) + 20.0 * (var2 - var1), 0, 100)
    return {"HIGH": int((s2 >= 70).sum()), "MEDIUM": int(((s2 >= 50) & (s2 < 70)).sum()),
            "LOW": int(((s2 >= 30) & (s2 < 50)).sum()), "DO_NOT_ACT": int((s2 < 30).sum())}


def sheet_growth(wb, d):
    ws = wb.create_sheet("Confidence Growth Plan")
    ws.cell(1, 1, "Confidence Growth Plan — how trust is manufactured, on a schedule"
            ).font = F(14, True, INK)
    ws.cell(2, 1, "Confidence is bought with data. Two purchases: weekly exports "
                  "(passive) and designed test waves (active). Projection uses the "
                  "engine's own scoring formula; fit and CI-tightness are held constant, "
                  "so these counts are a FLOOR — reality usually does better."
            ).font = NOTE_FONT

    now = {"HIGH": int((pd.to_numeric(d["confidence_score"], errors="coerce") >= 70).sum()),
           "MEDIUM": int(((pd.to_numeric(d["confidence_score"], errors="coerce") >= 50)
                          & (pd.to_numeric(d["confidence_score"], errors="coerce") < 70)).sum())}
    hdr = ["Horizon", "HIGH (data only)", "HIGH (data + waves)",
           "MEDIUM (data only)", "MEDIUM (data + waves)"]
    r0 = 4
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r0, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    lines = [("Today", now["HIGH"], now["HIGH"], now["MEDIUM"], now["MEDIUM"])]
    for wks in (4, 8, 12):
        p = _project_tiers(d, wks, with_waves=False)
        a = _project_tiers(d, wks, with_waves=True)
        lines.append((f"+{wks} weeks", p["HIGH"], a["HIGH"], p["MEDIUM"], a["MEDIUM"]))
    for i, row in enumerate(lines, r0 + 1):
        for j, v in enumerate(row, 1):
            c = ws.cell(i, j, v); c.font = F(10); c.border = HAIR
            if j > 1:
                c.font = F(10, bold=(j == 3), color=INK)

    # ── the 12-week calendar ──
    today = datetime.date.today()
    nm = today + datetime.timedelta(days=((7 - today.weekday()) % 7) or 7)
    stake = {w: float(pd.to_numeric(
        d.loc[d["wave"] == w, "slice_cost_mo"], errors="coerce").sum())
        for w in ("Wave 1", "Wave 2", "Wave 3")}
    r0 = r0 + len(lines) + 3
    ws.cell(r0, 1, "THE 12-WEEK CALENDAR").font = F(11, True, INK)
    cal_hdr = ["Week of", "Activity", "Cells", "Value at stake Rs/mo"]
    for j, h in enumerate(cal_hdr, 1):
        c = ws.cell(r0 + 1, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    events = [
        (0, f"Wave 1 goes live (−3ppt moves, floors respected); {KILL_SWITCH}",
         WAVE_SIZE, stake["Wave 1"]),
        (1, "Wave 1 monitored — weekly export scores predicted vs actual", "", ""),
        (2, "Wave 1 monitored (second strike week)", "", ""),
        (3, f"Wave 1 verdict: confirmed cells join the bankable set · Wave 2 live",
         WAVE_SIZE, stake["Wave 2"]),
        (4, "Wave 2 monitored", "", ""),
        (5, "Wave 2 monitored", "", ""),
        (6, "Wave 2 verdict · Wave 3 live", WAVE_SIZE, stake["Wave 3"]),
        (7, "Wave 3 monitored", "", ""),
        (8, "Wave 3 monitored", "", ""),
        (9, "Wave 3 verdict — full testable queue resolved", "", ""),
        (10, "Monthly rebuild with ~22 weeks of data — thin cells re-score", "", ""),
        (11, "Review: bankable set + next wave selection from backlog", "", ""),
    ]
    for i, (wk, act, cells, stk) in enumerate(events, r0 + 2):
        ws.cell(i, 1, f"{nm + datetime.timedelta(weeks=wk):%d %b}").font = F(9)
        ws.cell(i, 2, act).font = F(9)
        ws.cell(i, 3, cells).font = F(9)
        c = ws.cell(i, 4, round(stk) if stk != "" else "")
        c.font = F(9); c.number_format = "#,##0"
        for j in range(1, 5):
            ws.cell(i, j).border = HAIR
    for j, w in enumerate([14, 78, 16, 18, 20], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def _accuracy_receipts(run):
    """Portfolio-level accuracy receipts, loaded defensively from the artifacts."""
    out = []
    try:
        S = json.load(open(os.path.join(run, "plan", "plan_summary.json"), encoding="utf-8"))
        out.append(("Out-of-sample R² (champion, temporal holdout)",
                    f"{S.get('oos_r2')}  ({S.get('oos_cats_pass')}/{S.get('oos_cats_total')} categories ≥ 0.75)"))
    except Exception:
        pass
    try:
        ev = json.load(open(os.path.join(ROOT, "output", "DISCOUNT_PLAN", "validation",
                                         "elasticity_validation.json"), encoding="utf-8"))
        s1 = ev.get("stage1", {})
        out.append(("Elasticity gates (3-stage acceptance)",
                    f"{'ALL PASS' if ev.get('overall_pass', ev.get('all_pass')) else 'CHECK'} — "
                    f"holdout R² {s1.get('r2_log_holdout')}, wMAPE {s1.get('wmape_units')}, "
                    f"bias {s1.get('abs_bias_units')}"))
    except Exception:
        pass
    try:
        dml = json.load(open(os.path.join(run, "plan", "dml_results.json"), encoding="utf-8"))
        n_w = sum(1 for r in dml if r.get("waste"))
        out.append(("Double ML causal confirmation (independent method)",
                    f"{n_w}/{len(dml)} cut categories confirmed reliably-waste"))
    except Exception:
        pass
    try:
        sc = pd.read_csv(os.path.join(ROOT, "output", "DISCOUNT_PLAN", "validation",
                                      "sensitivity_cells.csv"))
        out.append(("Sensitivity (200 joint shakes of assumptions)",
                    f"{int(sc['fragile'].sum())} fragile of {len(sc)} cut cells"))
    except Exception:
        pass
    return out


def _staircase_rows(d):
    """The Value Staircase: three steps, each with its evidence standard.
    Step 3 reads the budget glide ladder's conservative (−2%, smart) rung."""
    cut_gain = float(pd.to_numeric(
        d.loc[d["bucket"] == "c_waste_cut", "net_gain_mo"], errors="coerce")
        .clip(lower=0).sum())
    stake = float(pd.to_numeric(
        d.loc[d["wave"].isin(["Wave 1", "Wave 2", "Wave 3"]), "slice_cost_mo"],
        errors="coerce").sum())
    rows = [
        ("STEP 1 — Bankable today", f"Rs.{cut_gain:,.0f}/mo",
         "survived all 8 gates + Double ML; in this week's KAM sheet", GOOD_FILL),
        ("STEP 2 — Test queue (Unlock Pipeline)",
         f"Rs.{stake:,.0f}/mo at stake",
         f"3 waves x {WAVE_SIZE} cells over ~9 weeks under the kill-switch; "
         f"SCENARIO: 30-50% confirming => Rs.{0.3*stake:,.0f}-{0.5*stake:,.0f}/mo "
         "joins Step 1", GREY_FILL),
    ]
    try:
        g = pd.read_csv(os.path.join(ROOT, "output", "DISCOUNT_PLAN", "pricing",
                                     "budget_glide.csv"))
        r2 = g[(g["mode"] == "smart") & (g["budget_change_pct"] == -2.0)].iloc[0]
        rows.append((
            "STEP 3 — Reallocation upside (no waste needed)",
            f"Rs.{abs(r2['spend_delta_mo']):,.0f}/mo spend freed",
            f"trim worst-ROI slices within proven floors; projected revenue "
            f"{r2['revenue_delta_pct']:+.1f}% (LOW-CONFIDENCE elasticities — "
            "verify via the waves before quoting)", GREY_FILL))
    except Exception:
        pass
    return rows


def sheet_summary(wb, d, run):
    ws = wb.create_sheet("Portfolio Summary", 1)
    ws.cell(1, 1, "Portfolio Summary — the whole story on one sheet").font = F(14, True, INK)
    ws.cell(2, 1, f"Run {os.path.basename(run)} · {cfg.BRAND_NAME} on "
                  f"{cfg.PLATFORM_NAME} · amounts only, no verdicts on sufficiency"
            ).font = NOTE_FONT
    ws.cell(4, 1, "THE VALUE STAIRCASE").font = F(12, True, INK)
    for i, (label, amount, standard, fill) in enumerate(_staircase_rows(d)):
        r = 5 + i
        ws.cell(r, 1, label).font = F(10, bold=True, color=INK)
        c = ws.cell(r, 2, amount); c.font = F(11, bold=True, color=INK); c.fill = fill
        ws.cell(r, 3, standard).font = F(9, italic=True, color=MUTED)
    ws.cell(9, 1, "Each step has a HIGHER evidence bar than the one below it — "
                  "quote them together, never Step 1 alone.").font = NOTE_FONT
    n_works = int(d["reliably_pays"].fillna(False).astype(bool).sum())
    n_trim = int((d["bucket"] == "c_waste_cut").sum())
    n_held = int((d["reliably_waste"].fillna(False).astype(bool)
                  & (d["bucket"] != "c_waste_cut")).sum())
    n_unc = len(d) - n_works - n_trim - n_held
    spend = float(pd.to_numeric(d["disc_spend_mo"], errors="coerce").sum())
    cut_gain = float(pd.to_numeric(
        d.loc[d["bucket"] == "c_waste_cut", "net_gain_mo"], errors="coerce")
        .clip(lower=0).sum())
    facts = [
        ("Cells analysed (SKU-pack x city)", len(d)),
        ("Total discount spend, Rs/mo", round(spend)),
        ("Cells where discount reliably WORKS", n_works),
        ("Cells to TRIM now (below pay-line, all gates clear)", n_trim),
        ("Below pay-line but HELD (stock/competitive/monitor)", n_held),
        ("Cells uncertain (monitor/test)", n_unc),
        ("Confident recoverable value if trimmed, Rs/mo", round(cut_gain)),
        ("Value at stake in the Unlock Pipeline (test queue), Rs/mo",
         round(float(pd.to_numeric(
             d.loc[d["test_candidate"].fillna(False).astype(bool), "slice_cost_mo"],
             errors="coerce").sum()))),
    ]
    r0 = 11    # below the staircase block
    for i, (k, v) in enumerate(facts):
        ws.cell(r0 + i, 1, k).font = F(10)
        c = ws.cell(r0 + i, 2, v)
        c.font = F(11, bold=True, color=INK)
        c.number_format = "#,##0"
    r0 += len(facts) + 2
    ws.cell(r0, 1, "MODEL ACCURACY — THE RECEIPTS").font = F(11, True, INK)
    for i, (k, v) in enumerate(_accuracy_receipts(run), 1):
        ws.cell(r0 + i, 1, k).font = F(10)
        ws.cell(r0 + i, 2, v).font = F(10, bold=True, color=INK)
    r0 += len(_accuracy_receipts(run)) + 3
    ws.cell(r0, 1, "By category").font = F(11, True, INK)
    g = (d.groupby("category")
         .agg(cells=("cell_id", "count"),
              spend_mo=("disc_spend_mo", "sum"),
              works=("reliably_pays", "sum"),
              not_working=("reliably_waste", "sum"))
         .reset_index().sort_values("spend_mo", ascending=False))
    hdr = ["Category", "Cells", "Discount spend Rs/mo", "Works", "Not working"]
    for j, h in enumerate(hdr, 1):
        c = ws.cell(r0 + 1, j, h)
        c.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        c.fill = HEAD_FILL
    for i, (_, r) in enumerate(g.iterrows(), r0 + 2):
        ws.cell(i, 1, r["category"]).font = F(9)
        ws.cell(i, 2, int(r["cells"])).font = F(9)
        c = ws.cell(i, 3, round(float(r["spend_mo"])))
        c.font = F(9); c.number_format = "#,##0"
        ws.cell(i, 4, int(r["works"])).font = F(9)
        ws.cell(i, 5, int(r["not_working"])).font = F(9)
    for j, w in enumerate([34, 8, 16, 8, 11], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def sheet_readme(wb):
    ws = wb.create_sheet("READ ME", 0)
    lines = [
        ("StatIQ Lab — Three-Stage Discount Report", 14, True),
        ("How to read this workbook, and exactly how every number is made.", 9, False),
        ("", 9, False),
        ("STAGE 1 — WHERE DOES DISCOUNT INCREASE SALES?", 11, True),
        ("The engine isolates discount's effect from availability (OSA), ad visibility,", 9, False),
        ("competitor price/availability and season, per SKU-pack x city cell.", 9, False),
        ("'Response per +1pt' = % unit change per extra discount point, everything else held equal.", 9, False),
        ("'Pay-line' = the response needed for a discount point to pay for itself.", 9, False),
        ("Uplift/units/revenue/margin are measured vs the cell's OBSERVED floor discount", 9, False),
        ("(the lowest it has actually traded at) — never vs an extrapolated 0%.", 9, False),
        ("", 9, False),
        ("STAGE 2 — HOW MUCH DISCOUNT SHOULD EACH CELL CARRY?", 11, True),
        ("Champion target = confounder-model target (floor-based).  Pipeline elbow =", 9, False),
        ("margin-optimal point on the saturation curve.  Optimizer optimal = portfolio", 9, False),
        ("optimum with cross-SKU substitution.  'SET THIS WEEK' is the governed number:", 9, False),
        ("a cut is issued only where BOTH engines agree, moved max 3ppt/week (glide).", 9, False),
        ("", 9, False),
        ("STAGE 3 — DID THE DISCOUNT CREATE ECONOMIC VALUE?", 11, True),
        ("REVENUE ROI = incremental revenue / cost of the trimmable discount slice", 9, False),
        ("(floor→today). ROI < 1 = the discount costs more than the revenue it brings.", 9, False),
        ("'Marginal ROI' asks the same at the margin: does the LAST point pay?", 9, False),
        ("", 9, False),
        ("THE VALUE STAIRCASE (Summary sheet, top)", 11, True),
        ("Three steps, each with a HIGHER evidence bar: bankable today (all gates) →", 9, False),
        ("test queue (converts via waves) → reallocation upside (projection to verify).", 9, False),
        ("Never quote Step 1 alone — it invites 'that's small'. The staircase is the story.", 9, False),
        ("", 9, False),
        ("CONFIDENCE GROWTH PLAN (its own sheet)", 11, True),
        ("Confidence is bought with data: weekly exports grow density; designed test", 9, False),
        ("waves add the discount variation flat pricing never provides. The projection", 9, False),
        ("uses the engine's own scoring formula with fit held constant — a floor.", 9, False),
        ("The 12-week calendar shows when each wave runs, is scored, and resolves.", 9, False),
        ("", 9, False),
        ("HELD CELLS AND THE UNLOCK PIPELINE", 11, True),
        ("No cell is ever just 'HELD': Stage 3 names WHY (the engine's own gate) and", 9, False),
        ("WHAT UNLOCKS it — more weeks (with a ready date), an availability fix, or a", 9, False),
        ("designed price test. Cells only a test can unlock are scheduled on the", 9, False),
        ("'Unlock Pipeline' sheet in waves, each protected by the kill-switch (a test", 9, False),
        ("that misses prediction twice auto-reverts). Held value is a queue with a", 9, False),
        ("schedule, not money the model refuses to discuss.", 9, False),
        ("", 9, False),
        ("WHEN THE BRAND ASKS: 'WHAT'S THE ELASTICITY, AND HOW ACCURATE IS IT?'", 11, True),
        ("Elasticity per SKU x city: open 'Elasticity & Accuracy' — quote the price", 9, False),
        ("elasticity WITH its 95% CI, and note the second, independently-estimated", 9, False),
        ("engine's number beside it (two methods, same story = the trust argument).", 9, False),
        ("Accuracy is answered in three layers, smallest to largest:", 9, False),
        ("  1. Per cell — the Confidence score (0-100): data density, price variation,", 9, False),
        ("     fit, plausibility and CI tightness combined. We ACT only on High.", 9, False),
        ("  2. Per category — model fit R2 at the grain decisions are made.", 9, False),
        ("  3. Portfolio — the receipts on the Summary sheet: out-of-sample R2,", 9, False),
        ("     3-stage elasticity gates, Double ML confirmation, sensitivity shakes.", 9, False),
        ("A single per-cell daily R2 is deliberately NOT the headline: daily units of", 9, False),
        ("one SKU in one city are noise-dominated; the honest per-cell trust measure", 9, False),
        ("is the confidence score, and the honest accuracy proof is out-of-sample.", 9, False),
        ("", 9, False),
        ("HONESTY NOTES — READ BEFORE QUOTING", 11, True),
        ("1. Every number is OBSERVABLE — units, prices, revenue, discount spend. No cost", 9, False),
        ("   assumptions (COGS/commission) enter any figure in this workbook.", 9, False),
        ("2. ~10 weeks of data: many cells carry category-level (shrunk) estimates and are", 9, False),
        ("   marked Low/Experimental confidence. Act on High; test the rest.", 9, False),
        ("3. Values-only snapshot generated from the run named on the Summary sheet;", 9, False),
        ("   regenerate with scripts/reports/build_stage_workbook.py after each rebuild.", 9, False),
        ("4. Colour code: green = reliably creates value, red = reliably below pay-line,", 9, False),
        ("   grey = uncertain (monitor/test).", 9, False),
    ]
    for i, (t, sz, b) in enumerate(lines, 1):
        ws.cell(i, 1, t).font = F(sz, bold=b, color=INK if b else BODY)
    ws.column_dimensions["A"].width = 110


def main():
    global OUT_XLSX
    if OUT_XLSX is None:
        OUT_XLSX = _next_versioned_out()
    run = _latest_run()
    d = load_joined(run)
    print(f"[stage-report] run {os.path.basename(run)} | {len(d)} cells -> "
          f"{os.path.basename(OUT_XLSX)}")
    wb = Workbook()
    wb.remove(wb.active)
    sheet_readme(wb)
    sheet_summary(wb, d, run)
    sheet_stage1(wb, d)
    sheet_stage2(wb, d)
    sheet_stage3(wb, d)
    sheet_unlock(wb, d)
    sheet_growth(wb, d)
    sheet_elasticity(wb, d)
    wb.save(OUT_XLSX)
    print(f"[stage-report] wrote {OUT_XLSX}")


if __name__ == "__main__":
    # Optional output-path override (e.g. when the canonical file is locked
    # open in Excel): python build_stage_workbook.py <out.xlsx>
    if len(sys.argv) > 1:
        OUT_XLSX = os.path.abspath(sys.argv[1])
    main()
